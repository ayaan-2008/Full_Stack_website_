from django.contrib import admin
from django.contrib.admin import AdminSite
from django.http import HttpResponse
from openpyxl import Workbook
import io
from .models import (
    SiteSettings, NavigationItem, FooterLinkGroup, FooterLink, SearchedTerm,
    Technology, Course, Branch, Programme, Testimonial, SuccessStory,
    HiringPartner, Certification, Blog, GalleryImage,
    Enquiry, CallbackRequest, RecruiterContact, BrochureRequest, CourseBrochure,
    LiveProject, Internship,
)


class ATAdminSite(AdminSite):
    site_header = 'AT Academy Admin'
    site_title = 'AT Academy'
    index_title = 'Dashboard'


admin_site = ATAdminSite(name='atacademy_admin')


def _export_xlsx(title, headers, rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = cell.font.copy(bold=True)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(val) if val else '')
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def export_selected_enquiries(modeladmin, request, queryset):
    headers = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Qualification', 'Date', 'Read']
    rows = [[e.name, e.email, e.phone, e.course, e.branch, e.qualification,
             e.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if e.is_read else 'No'] for e in queryset.order_by('-created_at')]
    return _export_xlsx('Enquiries', headers, rows, 'enquiries.xlsx')
export_selected_enquiries.short_description = 'Export selected as Excel'


def export_selected_callbacks(modeladmin, request, queryset):
    headers = ['Name', 'Email', 'Phone', 'Course', 'Branch', 'Date', 'Read']
    rows = [[c.name, c.email, c.phone, c.course, c.branch,
             c.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if c.is_read else 'No'] for c in queryset.order_by('-created_at')]
    return _export_xlsx('Callback Requests', headers, rows, 'callbacks.xlsx')
export_selected_callbacks.short_description = 'Export selected as Excel'


def export_selected_recruiters(modeladmin, request, queryset):
    headers = ['Name', 'Email', 'Phone', 'Company', 'Designation', 'Date', 'Read']
    rows = [[r.name, r.email, r.phone, r.company_name, r.designation,
             r.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if r.is_read else 'No'] for r in queryset.order_by('-created_at')]
    return _export_xlsx('Recruiter Contacts', headers, rows, 'recruiters.xlsx')
export_selected_recruiters.short_description = 'Export selected as Excel'


def export_selected_brochures(modeladmin, request, queryset):
    headers = ['Name', 'Email', 'Phone', 'Course', 'Date', 'Read']
    rows = [[b.name, b.email, b.phone, b.course.name if b.course else '',
             b.created_at.strftime('%Y-%m-%d %H:%M'), 'Yes' if b.is_read else 'No'] for b in queryset.select_related('course').order_by('-created_at')]
    return _export_xlsx('Brochure Requests', headers, rows, 'brochures.xlsx')
export_selected_brochures.short_description = 'Export selected as Excel'


@admin.register(SiteSettings, site=admin_site)
class SiteSettingsAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        return not SiteSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    fieldsets = (
        ('Basic Info', {'fields': ('site_name', 'tagline')}),
        ('Footer', {'fields': ('phone', 'email', 'address', 'google_maps_url'), 'description': 'These values appear in the footer section of the website.'}),
        ('Home Page Settings', {'fields': ('career_guidance_icon', 'materials_icon', 'interview_prep_icon', 'placement_icon', 'excel_banner'), 'description': 'Upload images for the "Excel with AT Academy" section on the homepage.'}),
        ('Nutshell / Career Services', {'fields': ('nutshell_illustration', 'nutshell_1', 'nutshell_2', 'nutshell_3', 'nutshell_4', 'nutshell_5', 'nutshell_6'), 'description': 'Upload images for the "Career Services in a Nutshell" section on the homepage.'}),
        ('Links', {'fields': ('student_portal_url', 'videos_url')}),
        ('Social Media', {'fields': ('social_media',)}),
        ('Branding', {'fields': ('header_logo', 'footer_logo')}),
    )


@admin.register(NavigationItem, site=admin_site)
class NavigationItemAdmin(admin.ModelAdmin):
    list_display = ['label', 'href', 'order', 'has_dropdown']
    list_editable = ['order', 'has_dropdown']
    ordering = ['order']


@admin.register(FooterLinkGroup, site=admin_site)
class FooterLinkGroupAdmin(admin.ModelAdmin):
    list_display = ['title', 'order']
    list_editable = ['order']


@admin.register(FooterLink, site=admin_site)
class FooterLinkAdmin(admin.ModelAdmin):
    list_display = ['label', 'group', 'href', 'order']
    list_editable = ['order']
    list_filter = ['group']


@admin.register(SearchedTerm, site=admin_site)
class SearchedTermAdmin(admin.ModelAdmin):
    list_display = ['label', 'href', 'order']
    list_editable = ['order']


@admin.register(Technology, site=admin_site)
class TechnologyAdmin(admin.ModelAdmin):
    list_display = ['name', 'order']
    list_editable = ['order']


class BrochureInline(admin.TabularInline):
    model = CourseBrochure
    extra = 1


@admin.register(Course, site=admin_site)
class CourseAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'fee', 'duration', 'order']
    list_editable = ['order']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name', 'category']
    list_filter = ['category']
    fields = ['name', 'slug', 'fee', 'duration', 'image', 'image_file', 'description', 'features', 'category', 'order', 'meta_title', 'meta_description', 'technologies']
    inlines = [BrochureInline]


@admin.register(Branch, site=admin_site)
class BranchAdmin(admin.ModelAdmin):
    list_display = ['name', 'address', 'phone', 'order']
    list_editable = ['order']
    prepopulated_fields = {'slug': ('name',)}
    search_fields = ['name']


@admin.register(Programme, site=admin_site)
class ProgrammeAdmin(admin.ModelAdmin):
    list_display = ['course', 'institution_name', 'mode', 'duration', 'fee_range', 'order']
    list_editable = ['order']
    list_filter = ['course', 'mode']


@admin.register(Testimonial, site=admin_site)
class TestimonialAdmin(admin.ModelAdmin):
    list_display = ['name', 'course', 'order']
    list_editable = ['order']


@admin.register(SuccessStory, site=admin_site)
class SuccessStoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'course', 'order']
    list_editable = ['order']


@admin.register(LiveProject, site=admin_site)
class LiveProjectAdmin(admin.ModelAdmin):
    list_display = ['title', 'client', 'github_url', 'order']
    list_editable = ['order']


@admin.register(Internship, site=admin_site)
class InternshipAdmin(admin.ModelAdmin):
    list_display = ['name', 'role', 'company', 'order']
    list_editable = ['order']


@admin.register(HiringPartner, site=admin_site)
class HiringPartnerAdmin(admin.ModelAdmin):
    list_display = ['name', 'order']
    list_editable = ['order']


@admin.register(Certification, site=admin_site)
class CertificationAdmin(admin.ModelAdmin):
    list_display = ['name', 'order']
    list_editable = ['order']


@admin.register(Blog, site=admin_site)
class BlogAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'date', 'order']
    list_editable = ['order']
    prepopulated_fields = {'slug': ('title',)}
    list_filter = ['category', 'date']
    search_fields = ['title']


@admin.register(GalleryImage, site=admin_site)
class GalleryImageAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'order']
    list_editable = ['order']
    list_filter = ['category']


@admin.register(Enquiry, site=admin_site)
class EnquiryAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone', 'course', 'branch', 'created_at', 'is_read']
    list_editable = ['is_read']
    list_filter = ['is_read', 'created_at', 'course', 'branch']
    search_fields = ['name', 'email', 'phone']
    readonly_fields = ['name', 'email', 'phone', 'course', 'branch', 'qualification', 'created_at']
    actions = [export_selected_enquiries]


@admin.register(CallbackRequest, site=admin_site)
class CallbackRequestAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone', 'course', 'branch', 'created_at', 'is_read']
    list_editable = ['is_read']
    list_filter = ['is_read', 'created_at']
    search_fields = ['name', 'email', 'phone']
    readonly_fields = ['name', 'email', 'phone', 'course', 'branch', 'created_at']
    actions = [export_selected_callbacks]


@admin.register(RecruiterContact, site=admin_site)
class RecruiterContactAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone', 'company_name', 'designation', 'created_at', 'is_read']
    list_editable = ['is_read']
    list_filter = ['is_read', 'created_at']
    search_fields = ['name', 'email', 'phone']
    readonly_fields = ['name', 'email', 'phone', 'company_name', 'designation', 'created_at']
    actions = [export_selected_recruiters]


@admin.register(BrochureRequest, site=admin_site)
class BrochureRequestAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'phone', 'course', 'created_at', 'is_read']
    list_editable = ['is_read']
    list_filter = ['is_read', 'created_at', 'course']
    search_fields = ['name', 'email', 'phone']
    readonly_fields = ['name', 'email', 'phone', 'course', 'created_at']
    actions = [export_selected_brochures]
